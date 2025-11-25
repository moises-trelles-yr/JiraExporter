import requests
import pandas as pd
import openpyxl
from datetime import datetime
import os
import json
from requests.auth import HTTPBasicAuth

# ==========================================
# CONFIGURATION
# ==========================================

# Jira Connection Details
JIRA_URL = "https://groupe-rocher.atlassian.net"  # Replace with your Jira instance URL
JIRA_API_EMAIL = "moises.trelles@yrnet.com"  # Replace with your email
JIRA_API_TOKEN = "ATATT3xFfGF0pN72Sxcz5ebzM_BClp0Mf3l9LbZx8v4v7W9g13kBr0rcZ8-ivbWeiP3kWIkGHALY0ExpLFydMStK2BdTyDRh6XEExxz5uCxXvGFDlztK7GD4tQubE18ZRp0czWrJnw5gmP8SSf-GMkJJijX6dialt1lUEvmT4KvRWbd5mvWhvWo=A03BD7F3"  # Replace with your API token

# JQL Query
JQL_QUERY = "filter=12408"  # Replace with your JQL

# Field Mapping
# IMPORTANT: You must replace the values below with the actual Field IDs from your Jira instance.
# Custom fields usually look like 'customfield_10001'.
# Standard fields like 'key', 'summary', 'status' are already correct.
TARGET_FIELDS_MAPPING = {

    "key": "key",  # Standard Field
    "Brand": "customfield_10124",
    "Country": "customfield_10107",
    "Summary" : "summary", # Standard Field
    "Status" : "customfield_10355",
    "Product": "customfield_10406",
    "PI": "customfield_10463",
    "Start date": "customfield_10284",
    "Due Date" : "duedate", # Standard Field
    "PI Priority": "customfield_10405",
    "Project_name": "customfield_10467",
    "Product Owner": "customfield_10233",
    "Product Manager": "customfield_10231",
    "Tech Lead": "customfield_10470",
    "Central Tester": "customfield_10472",
    "Data Analyst": "customfield_10604",
    "SF Core": "customfield_10443",
    "SFMC": "customfield_10438",
    "GCP INT": "customfield_10426",
    "Tableau": "customfield_10433",
    "MyCon": "customfield_10457",
    "Legacy LCDA": "customfield_10448",
    "Legacy ISAM": "customfield_10449",
    "Legacy PREVYR": "customfield_10450",
    "Legacy MyREPORT": "customfield_10451",
    "ST – STL": "customfield_10416",
    "ST – Devops": "customfield_10417",
    "ST – Legacy": "customfield_10410",
    "ST – Cloud Team": "customfield_10418",
    "HIP": "customfield_10419",
    "JOY": "customfield_10421",
    "VPCI": "customfield_10422",
    "PLANET": "customfield_10420",
    "DIGITAL": "customfield_10423",
    "RETAIL": "customfield_10424",
    "SAP": "customfield_10425",
    "CPG PO": "customfield_10441",
    "CPG PM": "customfield_10442",
    "CPG TL": "customfield_10439",
    "CPG SA": "customfield_10440",
    "DF PO": "customfield_10429",
    "DF PM": "customfield_10430",
    "DF TL": "customfield_10427",
    "DF SA": "customfield_10428",
    "DF DA": "customfield_10605",
    "DF REC CENTRAL": "customfield_10431",
    "DF REC LOCAL": "customfield_10432",
    "BI PO": "customfield_10436",
    "BI PM": "customfield_10437",
    "BI TL": "customfield_10434",
    "BI SA": "customfield_10435",
    "CC PO": "customfield_10446",
    "CC PM": "customfield_10447",
    "CC TL": "customfield_10444",
    "CC SA": "customfield_10445",
    "LEG PO": "customfield_10454",
    "LEG PM": "customfield_10455",
    "LEG TL": "customfield_10452",
    "LEG SA": "customfield_10453",
    "COE PO": "customfield_10505",
    "COE TL": "customfield_10506",


}

# Excel Export Configuration
EXCEL_FILENAME = "jira_export.xlsx"  # The Excel file to update
EXCEL_SHEET_NAME = "Features"  # The sheet name to update

# ==========================================
# FUNCTIONS
# ==========================================

def get_jira_headers():
    """Returns the headers for Jira API authentication."""
    return {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }


def search_jira_tickets(jql, fields_mapping):
    """
    Searches for Jira tickets using the provided JQL and fields.
    Uses Jira API v3 /rest/api/3/search/jql endpoint with nextPageToken pagination.
    Handles pagination to fetch all results.
    """
    url = f"{JIRA_URL}/rest/api/3/search/jql"
    auth = HTTPBasicAuth(JIRA_API_EMAIL, JIRA_API_TOKEN)
    headers = get_jira_headers()

    # Extract the list of field IDs to request from Jira
    fields_to_request = list(fields_mapping.values())

    max_results = 100
    all_issues = []
    next_page_token = None

    print(f"Starting Jira search with JQL: {jql}")

    while True:
        params = {
            "jql": jql,
            "maxResults": max_results,
            "fields": ",".join(fields_to_request)
        }

        # Add nextPageToken only if it exists (not for the first page)
        if next_page_token:
            params["nextPageToken"] = next_page_token

        try:
            response = requests.get(url, headers=headers, params=params, auth=auth)
            response.raise_for_status()
            data = response.json()

            issues = data.get("issues", [])
            all_issues.extend(issues)

            print(f"Fetched {len(issues)} issues (Total: {len(all_issues)})")

            # Check if there's a next page
            next_page_token = data.get("nextPageToken")

            # If nextPageToken is not in the response, we've reached the last page
            if "nextPageToken" not in data:
                break

        except requests.exceptions.RequestException as e:
            print(f"Error fetching data from Jira: {e}")
            if hasattr(e, 'response') and e.response is not None:
                print(f"Response content: {e.response.text}")
            break

    return all_issues


def get_field_value(issue, field_id):
    """
    Helper to extract value from a field, handling nested objects.
    """
    # Top level fields like 'key' are directly in the issue dict
    if field_id == 'key':
        return issue.get('key')

    fields = issue.get('fields', {})
    value = fields.get(field_id)

    if value is None:
        return None

    # Handle common nested structures
    if isinstance(value, dict):
        if 'name' in value:
            return value['name']
        elif 'displayName' in value:
            return value['displayName']
        elif 'value' in value:
            return value['value']
        elif 'summary' in value:  # For parent link etc
            return value['summary']

    # Handle lists (e.g. labels, multi-select)
    if isinstance(value, list):
        return ", ".join([str(v.get('value', v.get('name', v))) if isinstance(v, dict) else str(v) for v in value])

    return value


def process_tickets(tickets_data, fields_mapping):
    """
    Transforms the raw JSON data into a clean list of dictionaries.
    """
    processed_data = []

    for issue in tickets_data:
        row = {}
        for friendly_name, field_id in fields_mapping.items():
            row[friendly_name] = get_field_value(issue, field_id)
        processed_data.append(row)

    return pd.DataFrame(processed_data)


def export_to_excel(dataframe, filename, sheet_name="Features"):
    """
    Updates data in an existing Excel file while preserving:
    - Header formatting
    - Other columns not in the dataframe
    - Formulas in columns not being updated

    If the file doesn't exist, creates a new one.

    Args:
        dataframe: The DataFrame to export
        filename: The Excel file path to update
        sheet_name: The name of the sheet to update (default: "Features")

    Returns:
        The filename if successful, None otherwise
    """
    try:
        # Check if file exists
        if os.path.exists(filename):
            print(f"Updating existing file: {filename}")

            # Load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(filename)

            # Check if sheet exists
            if sheet_name not in wb.sheetnames:
                print(f"Sheet '{sheet_name}' not found. Creating new sheet.")
                ws = wb.create_sheet(sheet_name)
                # Write headers
                for col_idx, col_name in enumerate(dataframe.columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                header_row = 1
                column_mapping = {col_name: col_idx for col_idx, col_name in enumerate(dataframe.columns, start=1)}
            else:
                ws = wb[sheet_name]

                # Read existing headers from first row
                header_row = 1
                existing_headers = {}
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    if cell.value:
                        existing_headers[cell.value] = col_idx

                # Map dataframe columns to existing columns
                column_mapping = {}
                for col_name in dataframe.columns:
                    if col_name in existing_headers:
                        column_mapping[col_name] = existing_headers[col_name]
                    else:
                        # Add new column at the end
                        max_col = ws.max_column + 1
                        ws.cell(row=header_row, column=max_col, value=col_name)
                        column_mapping[col_name] = max_col
                        print(f"Added new column: {col_name}")

                # Clear only the data rows for columns we're updating (preserve formulas by checking)
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    for col_name, col_idx in column_mapping.items():
                        cell = ws.cell(row=row_idx, column=col_idx)
                        # Only clear if it's not a formula
                        if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                            cell.value = None

            # Write dataframe data starting from row after header
            for df_row_idx, row_data in dataframe.iterrows():
                excel_row_idx = header_row + 1 + df_row_idx
                for col_name, col_idx in column_mapping.items():
                    value = row_data[col_name]
                    # Convert NaN/None to empty string
                    if pd.isna(value):
                        value = None
                    ws.cell(row=excel_row_idx, column=col_idx, value=value)

            # Save workbook
            wb.save(filename)
            wb.close()
            print(f"Successfully updated sheet '{sheet_name}' in {filename}")
            print(f"Updated {len(dataframe)} rows across {len(column_mapping)} columns")

        else:
            print(f"Creating new file: {filename}")
            # Create new file using pandas
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Successfully created {filename} with sheet '{sheet_name}'")

        return filename
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# ==========================================
# MAIN EXECUTION
# ==========================================

def main():
    print("Starting Jira Data Extraction...")

    # 1. Search Tickets
    tickets = search_jira_tickets(JQL_QUERY, TARGET_FIELDS_MAPPING)

    if not tickets:
        print("No tickets found or error occurred.")
        return

    # 2. Process Data
    df = process_tickets(tickets, TARGET_FIELDS_MAPPING)
    print(f"Processed {len(df)} records.")
    print(df.head())

    # 3. Export to Excel
    excel_file = export_to_excel(df, EXCEL_FILENAME, EXCEL_SHEET_NAME)


if __name__ == "__main__":
    main()
